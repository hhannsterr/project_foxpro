import os
import warnings

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

class process:
    def __init__(self, sheet):
        self.date = None
        self.products = []
        self.good_pcs = []
        self.sheet = sheet
        self.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.font = Font(name='Tahoma', size=8)
        self.format = '0'

    def get_info(self, report: str) -> None:
        items = report.strip().split('\n')

        for item in items:
            product, good_pc = item.strip().split()
            self.products.append(product)
            self.good_pcs.append(good_pc)

    def fetch_summary(self, path: str, filename: str) -> None:

        full_path = path + '/' + filename
        with open(full_path, "r") as file:
            content = file.read()
            content = content.replace('\x1a', '')
            self.date, content = get_date(content)

            if not is_mould(filename):
                self.get_info(content)

            else:
                reports = content.split('---')
                
                old_report = reports[1]
                new_report = reports[3]

                if old_report != '\n':
                    self.get_info(old_report)
                if new_report != '\n':
                    self.get_info(new_report)

    def populate_data(self, columns):
        id_col, pc_col = columns
            
        row_num = find_last_blank_row(self.sheet, id_col)

        cell = self.sheet[f'{id_col}{row_num}']
        cell.value = self.date
        cell.fill = self.fill
        cell.font = self.font

        row_num += 1

        for product, good_pc in zip(self.products, self.good_pcs):
            cell = self.sheet[f'{id_col}{row_num}']
            cell.value = product
            cell.font = self.font
            cell.number_format = self.format

            cell = self.sheet[f'{pc_col}{row_num}']
            cell.value = good_pc
            cell.font = self.font
            cell.number_format = self.format

            row_num += 1


def is_mould(filename: str) -> bool:
    return filename == 'MOULD_SUMMARY.TXT'

def get_date(content: str) -> tuple[str]:
    date, content = content.split('-----')
    return date.strip(), content.strip()

def load_excel(excel_path: str, sheet_name: str):
    workbook = load_workbook(excel_path)

    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Sheet '{sheet_name}' not found.")

    return workbook, workbook[sheet_name]

def clean_path(path):
    return path.split('_')[0].lower()

def find_last_blank_row(sheet, column_letter):
    row_num = 1
    while sheet[f"{column_letter}{row_num}"].value is not None:
        row_num += 1
    return row_num


def main():

    path_to_data = 'data'
    process_col = {
        'tap': ['F', 'G'],
        'mould': ['M', 'N'],
        'melt': ['Q', 'R'],
        'grind': ['U', 'V'],
        'sht': ['Y', 'Z'],
        'galv': ['AC', 'AD'],
    }

    excel_path = r'C:\Users\Admin\Desktop\Master Plan 16-4-2025.xlsx'
    sheet_name = 'data of Foxpro'

    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
    workbook, sheet = load_excel(excel_path, sheet_name)

    process_paths = os.listdir(path_to_data)
    for process_path in process_paths:
        temp = process(sheet)
        temp.fetch_summary(path_to_data, process_path)

        columns = process_col[clean_path(process_path)]
        temp.populate_data(columns)

    workbook.save('test.xlsx')


if __name__ == '__main__':
    main()

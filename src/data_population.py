from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

class process:
    def __init__(self, sheet):
        self.date = None
        self.products = []
        self.good_pcs = []
        self.sheet = sheet
        self.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.font = Font(name='Tahoma', size=8)
        self.format = '#,##0'

    def get_info(self, report: str) -> None:
        items = report.strip().split('\n')

        for item in items:
            product, good_pc = item.strip().split()
            self.products.append(product)
            self.good_pcs.append(good_pc)

    def get_info_gal(self, report: str) -> None:
        items = report.strip().split('\n')
        bg_index, next_index = 2, 3

        for item in items:
            product, good_pc = item.strip().split()
            product = product[:bg_index] + 'b' + product[next_index:]
            self.products.append(product)
            self.good_pcs.append(good_pc)

    def fetch_summary(self, path: str, filename: str) -> None:

        full_path = path + '/' + filename
        with open(full_path, "r") as file:
            content = file.read()
            content = content.replace('\x1a', '')
            self.date, content = get_date(content)

            if is_gal(filename):
                self.get_info_gal(content)

            elif not is_mould(filename):
                self.get_info(content)

            else:
                reports = content.split('---')
                
                old_report = reports[1]
                new_report = reports[3]

                if old_report != '\n':
                    self.get_info(old_report)
                if new_report != '\n':
                    self.get_info(new_report)

    def populate_data(self, columns: list[str]) -> None:
        id_col, pc_col = columns
            
        row_num = find_last_blank_row(self.sheet, id_col)

        cell = self.sheet[f'{id_col}{row_num}']
        cell.value = self.date
        cell.fill = self.fill
        cell.font = self.font

        row_num += 1

        for product, good_pc in zip(self.products, self.good_pcs):
            cell = self.sheet[f'{id_col}{row_num}']
            cell.value = product
            cell.font = self.font

            cell = self.sheet[f'{pc_col}{row_num}']
            cell.value = int(good_pc.replace(',', ''))
            cell.font = self.font
            cell.number_format = self.format

            row_num += 1


def is_gal(filename: str) -> bool:
    return filename == 'GALV_SUMMARY.TXT'

def is_mould(filename: str) -> bool:
    return filename == 'MOULD_SUMMARY.TXT'

def get_date(content: str) -> tuple[str]:
    date, content = content.split('-----')
    return date.strip(), content.strip()

def is_monday() -> bool:
    return datetime.now().weekday() == 0

def load_excel(excel_path: str, sheet_name: str):
    workbook = load_workbook(excel_path)

    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Sheet '{sheet_name}' not found.")

    return workbook, workbook[sheet_name]

def clean_path(path: str) -> str:
    return path.split('_')[0].lower()

def find_last_blank_row(sheet, column_letter: str) -> int:
    row_num = 1
    while sheet[f"{column_letter}{row_num}"].value is not None:
        row_num += 1
    return row_num
